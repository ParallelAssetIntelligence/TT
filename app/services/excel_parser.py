import io
import logging
from openpyxl import Workbook, load_workbook
from app.models.lead import LeadRow, EnrichedData

logger = logging.getLogger(__name__)


def parse_excel(file_bytes: bytes) -> list[LeadRow]:
    """
    Parse any Excel file into rows. No fixed columns required.
    Reads whatever headers exist and stores each row as a dict.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")

    # Row 0 = headers (keep original casing for output)
    headers = [str(h).strip() if h else f"Column_{i}" for i, h in enumerate(rows[0])]

    leads = []
    for row in rows[1:]:
        # Skip completely empty rows
        if not any(cell for cell in row):
            continue

        data = {}
        for i, header in enumerate(headers):
            cell_val = row[i] if i < len(row) else None
            data[header] = str(cell_val).strip() if cell_val is not None else ""

        leads.append(LeadRow(headers=headers, data=data))

    logger.info(f"Parsed {len(leads)} rows with {len(headers)} columns: {headers}")
    return leads


def write_enriched_excel(leads: list[LeadRow], enrichments: list[EnrichedData]) -> bytes:
    """
    Write enriched Excel: original columns + appended smart columns.
    """
    if not leads:
        raise ValueError("No leads to write")

    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Leads"

    # Smart columns to append (kept in sync with row_enricher.ENRICHED_COLUMNS,
    # minus the "Enriched_" prefix to keep the /upload output friendly).
    smart_columns = [
        "Website",
        "LinkedIn",
        "LinkedIn_Headline",
        "LinkedIn_Summary",
        "LinkedIn_Company_Description",
        "Title",
        "Description",
        "Location",
        "Tenure_Months",
        "Tenure_Label",
        "Prior_Company_1",
        "Prior_Company_2",
        "Title_Qualifier",
        "Signal_Tag",
        "Script_Used",
        "Personalized_Opener",
    ]

    # Headers = original headers + smart columns
    original_headers = leads[0].headers
    all_headers = original_headers + smart_columns
    ws.append(all_headers)

    # Data rows
    for lead, enriched in zip(leads, enrichments):
        # Original data in original order
        original_values = [lead.data.get(h, "") for h in original_headers]
        # Smart data appended (order MUST match smart_columns)
        smart_values = [
            enriched.website,
            enriched.linkedin,
            enriched.linkedin_headline,
            enriched.linkedin_summary,
            enriched.linkedin_company_description,
            enriched.title,
            enriched.description,
            enriched.location,
            enriched.tenure_months,
            enriched.tenure_label,
            enriched.prior_company_1,
            enriched.prior_company_2,
            enriched.title_qualifier,
            enriched.signal_tag,
            enriched.script_used,
            enriched.personalized_opener,
        ]
        ws.append(original_values + smart_values)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
