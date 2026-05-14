import logging
from pathlib import Path
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _load_sheet(file_path: str):
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    return wb.active


def _headers(ws) -> list[str]:
    return [
        str(c.value).strip() if c.value is not None else f"Column_{i}"
        for i, c in enumerate(ws[1])
    ]


def _row_to_dict(row_cells, headers: list[str]) -> dict[str, str]:
    return {
        header: (
            str(row_cells[i].value).strip()
            if i < len(row_cells) and row_cells[i].value is not None
            else ""
        )
        for i, header in enumerate(headers)
    }


def read_row(file_path: str, row_number: int) -> dict:
    """Read a single Excel row by its row number (row 1 = header, row 2 = first lead)."""
    ws = _load_sheet(file_path)
    if row_number < 2 or row_number > ws.max_row:
        raise ValueError(
            f"row must be between 2 and {ws.max_row} (row 1 is the header)"
        )

    headers = _headers(ws)
    target_cells = ws[row_number]
    data = _row_to_dict(target_cells, headers)

    return {"row": row_number, "headers": headers, "data": data}


def read_all_rows(file_path: str, limit: int | None = None, offset: int = 0) -> dict:
    """
    Read all lead rows.
    - offset is 0-indexed over data rows (not Excel rows). offset=0 → starts at Excel row 2.
    - limit is optional; if None, return everything from offset onwards.
    Empty rows are skipped.
    """
    ws = _load_sheet(file_path)
    headers = _headers(ws)

    rows: list[dict] = []
    excel_row_number = 1
    data_index = 0
    for cells in ws.iter_rows(min_row=2):
        excel_row_number += 1
        if not any(c.value for c in cells):
            continue
        if data_index < offset:
            data_index += 1
            continue
        rows.append({
            "row": excel_row_number,
            "data": _row_to_dict(cells, headers),
        })
        data_index += 1
        if limit is not None and len(rows) >= limit:
            break

    return {
        "headers": headers,
        "count": len(rows),
        "offset": offset,
        "limit": limit,
        "rows": rows,
    }
