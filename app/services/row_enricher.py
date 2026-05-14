import time
import logging
from pathlib import Path
from openpyxl import load_workbook
from app.models.lead import LeadRow, EnrichedData
from app.services.serpapi_enricher import enrich_lead

logger = logging.getLogger(__name__)

ENRICHED_COLUMNS = [
    "Enriched_Website",
    "Enriched_LinkedIn",
    "Enriched_LinkedIn_Headline",
    "Enriched_LinkedIn_Summary",
    "Enriched_LinkedIn_Company_Description",
    "Enriched_Title",
    "Enriched_Description",
    "Enriched_Location",
]


def parse_row_range(rows: str) -> tuple[int, int]:
    """Parse a user-supplied range string like '2', '2-10', or '2 - 10'."""
    s = rows.strip().replace(" ", "")
    if not s:
        raise ValueError("rows is empty — use '2' or '2-10'")

    parts = s.split("-")
    try:
        if len(parts) == 1:
            n = int(parts[0])
            return n, n
        if len(parts) == 2:
            return int(parts[0]), int(parts[1])
    except ValueError:
        pass
    raise ValueError(f"Invalid range '{rows}'. Use a number (e.g. '2') or a range (e.g. '2-10').")


def enrich_row_range(file_path: str, start: int, end: int) -> dict:
    """
    Enrich rows [start, end] (inclusive) in `file_path` and save back to the same file.
    Opens and saves the workbook only once.
    """
    if start > end:
        raise ValueError(f"start row ({start}) must be <= end row ({end})")

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Fail fast if the file is locked (e.g. open in Excel) so we don't waste SerpAPI calls.
    # Excel creates a hidden lock file named '~$<filename>' next to the open file.
    lock_file = path.parent / f"~${path.name}"
    if lock_file.exists():
        raise PermissionError(
            f"'{path.name}' is open in another program (likely Excel). Close it and try again."
        )
    try:
        with open(path, "r+b"):
            pass
    except PermissionError:
        raise PermissionError(
            f"'{path.name}' is locked by another process. Close it and try again."
        )

    wb = load_workbook(filename=str(path))
    ws = wb.active

    if start < 2 or end > ws.max_row:
        raise ValueError(
            f"rows must be between 2 and {ws.max_row} (row 1 is the header)"
        )

    headers = [
        str(c.value).strip() if c.value is not None else f"Column_{i}"
        for i, c in enumerate(ws[1])
    ]
    for col_name in ENRICHED_COLUMNS:
        if col_name not in headers:
            new_idx = len(headers) + 1
            ws.cell(row=1, column=new_idx, value=col_name)
            headers.append(col_name)

    original_count = len(headers) - len(ENRICHED_COLUMNS)

    results: list[dict] = []
    errors: list[dict] = []
    rows_to_process = list(range(start, end + 1))

    for i, row_number in enumerate(rows_to_process):
        row_cells = ws[row_number]
        row_data = {
            header: (
                str(row_cells[idx].value).strip()
                if idx < len(row_cells) and row_cells[idx].value is not None
                else ""
            )
            for idx, header in enumerate(headers[:original_count])
        }

        if not any(row_data.values()):
            errors.append({"row": row_number, "error": "empty row"})
            continue

        try:
            lead = LeadRow(headers=headers[:original_count], data=row_data)
            enriched: EnrichedData = enrich_lead(lead)
        except Exception as e:
            logger.exception(f"Row {row_number} enrichment failed")
            errors.append({"row": row_number, "error": str(e)})
            continue

        values = {
            "Enriched_Website": enriched.website,
            "Enriched_LinkedIn": enriched.linkedin,
            "Enriched_LinkedIn_Headline": enriched.linkedin_headline,
            "Enriched_LinkedIn_Summary": enriched.linkedin_summary,
            "Enriched_LinkedIn_Company_Description": enriched.linkedin_company_description,
            "Enriched_Title": enriched.title,
            "Enriched_Description": enriched.description,
            "Enriched_Location": enriched.location,
        }
        for col_name, val in values.items():
            col_idx = headers.index(col_name) + 1
            ws.cell(row=row_number, column=col_idx, value=val or "")

        present = {k: v for k, v in row_data.items() if v}
        missing = [k for k, v in row_data.items() if not v]

        results.append({
            "row": row_number,
            "search_query": lead.search_text(),
            "present": present,
            "missing": missing,
            "original": row_data,
            "enriched": values,
        })

        if i < len(rows_to_process) - 1:
            time.sleep(1)

    wb.save(str(path))
    logger.info(
        f"Saved {path.name}: {len(results)} rows enriched, {len(errors)} failed"
    )

    return {
        "rows_processed": len(results),
        "rows_failed": len(errors),
        "results": results,
        "errors": errors,
    }
