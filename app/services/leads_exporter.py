"""Export enriched leads from the Supabase `leads` table to an xlsx blob.

Used by GET /leads/download (linked from the Teams notification card) so Matt
can grab a per-upload spreadsheet whose layout matches the original file he
uploaded — every original column preserved in its original order — with our
Enriched_* columns appended on the right. Mirrors how the legacy
write_enriched_excel() function shaped its output.
"""
import io
import logging
from typing import Any

from openpyxl import Workbook

from app.databaseconnection import supabase_manager

logger = logging.getLogger(__name__)

LEADS_TABLE = "leads"

# Columns appended after the original Excel columns. (label_in_xlsx, dotted_db_path)
# A path like "enrichment.foo" reads from the row's enrichment jsonb;
# anything else reads from the top-level row dict.
ENRICHMENT_COLUMNS: list[tuple[str, str]] = [
    ("Enriched_Website", "enrichment.website"),
    ("Enriched_LinkedIn", "enrichment.linkedin_url"),
    ("Enriched_LinkedIn_Headline", "enrichment.linkedin_headline"),
    ("Enriched_LinkedIn_Summary", "enrichment.linkedin_summary"),
    ("Enriched_LinkedIn_Company_Description", "enrichment.linkedin_company_description"),
    ("Enriched_Location", "enrichment.location"),
    ("Enriched_Description", "enrichment.description"),
    ("Enriched_Tenure_Months", "enrichment.tenure_months"),
    ("Enriched_Tenure_Label", "enrichment.tenure_label"),
    ("Enriched_Prior_Company_1", "enrichment.prior_company_1"),
    ("Enriched_Prior_Company_2", "enrichment.prior_company_2"),
    ("Enriched_Title_Qualifier", "title_qualifier"),
    ("Enriched_Signal_Tag", "signal_tag"),
    ("Enriched_Script_Used", "enrichment.script_used"),
    ("Enriched_Personalized_Opener", "enrichment.personalized_opener"),
    ("Enrichment_Status", "enrichment_status"),
    ("Enriched_At", "enriched_at"),
]

# Fallback header order for rows that have NO raw_excel_row (legacy rows
# inserted before the v3 migration). Uses the top-level columns so the
# download is never empty.
LEGACY_FALLBACK_COLUMNS = [
    "name", "company", "title", "phone", "email",
]


def fetch_leads(source_file: str | None = None, limit: int = 5000) -> list[dict[str, Any]]:
    """Pull rows from the leads table, optionally filtered by source_file."""
    client = supabase_manager.get_client()
    if not client:
        raise RuntimeError("Supabase client unavailable")

    query = client.table(LEADS_TABLE).select("*").order("id")
    if source_file:
        query = query.eq("source_file", source_file)
    return query.limit(limit).execute().data or []


def leads_to_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Build an xlsx whose left side mirrors the original upload and whose
    right side has the Enriched_* columns appended.

    Original column ordering is recovered by walking rows in DB order and
    collecting `raw_excel_row` keys in first-seen order. Any row missing
    a raw_excel_row contributes nothing to the header list — its values
    will appear blank under the raw columns, but its enrichment data
    still flows through on the right.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Leads"

    # 1. Discover the original Excel columns (preserving first-seen order).
    raw_columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in (row.get("raw_excel_row") or {}).keys():
            if k and k not in seen:
                seen.add(k)
                raw_columns.append(k)

    # Fallback for legacy rows (uploaded before v3 migration) — emit the
    # basic top-level columns so the file isn't empty for them.
    if not raw_columns:
        raw_columns = [c.title() for c in LEGACY_FALLBACK_COLUMNS]
        use_legacy_fallback = True
    else:
        use_legacy_fallback = False

    # 2. Header row = original columns + enrichment columns.
    enrichment_labels = [label for label, _ in ENRICHMENT_COLUMNS]
    ws.append(raw_columns + enrichment_labels)

    # 3. Data rows.
    for row in rows:
        raw = row.get("raw_excel_row") or {}
        enrichment = row.get("enrichment") or {}

        raw_values: list[Any] = []
        for col in raw_columns:
            if use_legacy_fallback:
                raw_values.append(row.get(col.lower(), ""))
            else:
                raw_values.append(raw.get(col, ""))

        enriched_values: list[Any] = []
        for _, key in ENRICHMENT_COLUMNS:
            if key.startswith("enrichment."):
                enriched_values.append(enrichment.get(key.split(".", 1)[1], ""))
            else:
                enriched_values.append(row.get(key, ""))

        ws.append([_coerce_cell(v) for v in (raw_values + enriched_values)])

    # 4. Light auto-fit so the sheet is readable.
    for i, _ in enumerate(raw_columns + enrichment_labels, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        max_len = max(
            (len(str(cell.value)) for cell in ws[col_letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce_cell(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return str(v)
    return v
